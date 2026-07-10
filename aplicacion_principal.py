import pandas as pd
import streamlit as st
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Gestión de Nóminas DS19", layout="centered")

# Crear sistema de navegación por pestañas
pestana_pis, pestana_etapas = st.tabs(["Consolidador PIS (Asistencias)", "Validador de Etapas (Vigencia)"])

with pestana_pis:
    # --- MÓDULO 1: CONSOLIDADOR DE ASISTENCIAS PIS ---
    if 'lista_archivos' not in st.session_state:
        st.session_state.lista_archivos = []
    if 'llave_subida' not in st.session_state:
        st.session_state.llave_subida = 0

    st.title("Generador Nómina final PIS")
    st.write("1. Sube los archivos excel.\n2. Usa las flechas para ordenarlos cronológicamente (Actividad 1 arriba, la más reciente abajo). \n3. El último archivo definirá la vigencia y los datos más actualizados.")

    archivos_nuevos = st.file_uploader(
        "Arrastra o selecciona las nóminas aquí (los archivos se moverán a la lista de abajo)", 
        type=['xlsx'], 
        accept_multiple_files=True,
        key=f"uploader_pis_{st.session_state.llave_subida}"
    )

    if archivos_nuevos:
        nombres_actuales = [f.name for f in st.session_state.lista_archivos]
        for archivo in archivos_nuevos:
            if archivo.name not in nombres_actuales:
                st.session_state.lista_archivos.append(archivo)
        st.session_state.llave_subida += 1
        st.rerun()

    # Panel interactivo de ordenamiento con flechas restauradas
    if st.session_state.lista_archivos:
        st.markdown("### Orden de procesamiento")
        
        for i, archivo in enumerate(st.session_state.lista_archivos):
            col_texto, col_arriba, col_abajo, col_eliminar = st.columns([6, 1, 1, 1])
            
            col_texto.write(f"**{i+1}.** {archivo.name}")
            
            if col_arriba.button("⬆", key=f"up_{i}", disabled=(i == 0)):
                st.session_state.lista_archivos[i], st.session_state.lista_archivos[i-1] = st.session_state.lista_archivos[i-1], st.session_state.lista_archivos[i]
                st.rerun()
                
            if col_abajo.button("⬇", key=f"down_{i}", disabled=(i == len(st.session_state.lista_archivos) - 1)):
                st.session_state.lista_archivos[i], st.session_state.lista_archivos[i+1] = st.session_state.lista_archivos[i+1], st.session_state.lista_archivos[i]
                st.rerun()
                
            if col_eliminar.button("✕", key=f"del_{i}"):
                st.session_state.lista_archivos.pop(i)
                st.rerun()

    st.divider()

    if st.button("Generar matriz PIS", type="primary"):
        if not st.session_state.lista_archivos:
            st.warning("No hay archivos en la lista para procesar.")
        else:
            columnas_objetivo = [
                "Región", "Comuna", "Tipo de Aplicación", "Estado Reserva", "Rut", "Dv", 
                "Primer Apellido", "Segundo Apellido", "Nombres", "Rango Precio Vivienda", 
                "Código Proyecto", "Nombre Proyecto", "Rut Entidad Desarrolladora", 
                "Nombre Entidad Desarrolladora", "Subsidio", "Línea de Proceso", "Llamado", 
                "Año", "Nombre Oferta del Llamado", "Fecha", "Usuario", "Familia Objetivo", 
                "N° de viviendas del proyecto", "N° de Resolución de asignación", 
                "Fecha de Resolución de asignación", "N° de Resolución de asignación DS49", 
                "Fecha de Resolución de asignación DS49", "Código de grupo DS49", 
                "Estado de pago", "TRAMO CSE", "Monto pagado (U.F.)", "Fecha de pago"
            ]
            
            datos_maestros = {}
            etiquetas_actividades = []
            ruts_ultima_nomina = set()
            total_archivos = len(st.session_state.lista_archivos)
            
            for numero_actividad, archivo_en_memoria in enumerate(st.session_state.lista_archivos, start=1):
                datos_crudos = pd.read_excel(archivo_en_memoria, header=None)
                indice_encabezado = 0
                for i in range(min(15, len(datos_crudos))):
                    fila_texto = datos_crudos.iloc[i].astype(str).str.upper()
                    if any("RUT" in valor for valor in fila_texto) and any("NOMBRES" in valor for valor in fila_texto):
                        indice_encabezado = i
                        break
                        
                datos_actividad = pd.read_excel(archivo_en_memoria, header=indice_encabezado)
                columnas_actuales = {str(col).strip().upper(): col for col in datos_actividad.columns}
                
                if 'RUT' not in columnas_actuales:
                    continue
                    
                columna_rut_real = columnas_actuales['RUT']
                etiqueta_asistencia = f"asistencia_actividad_{numero_actividad}"
                etiquetas_actividades.append(etiqueta_asistencia)
                
                tiene_control_asistencia = False
                columna_asistencia_real = None
                for palabra_clave in ['FIRMA', 'ASISTENCIA', 'ASISTENCIAS', 'CONTROL ASISTENCIA', 'FIRMAS', 'REGISTRO ASISTENCIA', 'REGISTRO DE ASISTENCIA', 'REGISTRO DE FIRMA', 'REGISTRO DE FIRMAS']:
                    if palabra_clave in columnas_actuales:
                        tiene_control_asistencia = True
                        columna_asistencia_real = columnas_actuales[palabra_clave]
                        break
                
                registros = datos_actividad.to_dict('records')
                for fila in registros:
                    rut_sucio = str(fila[columna_rut_real])
                    rut_limpio = ''.join(caracter for caracter in rut_sucio.upper() if caracter.isalnum())
                    
                    if len(rut_limpio) < 7:
                        continue
                    if numero_actividad == total_archivos:
                        ruts_ultima_nomina.add(rut_limpio)
                    if rut_limpio not in datos_maestros:
                        datos_maestros[rut_limpio] = {col: pd.NA for col in columnas_objetivo}
                        datos_maestros[rut_limpio]['rut_interno'] = rut_limpio
                    
                    for col_obj in columnas_objetivo:
                        col_obj_upper = col_obj.upper()
                        if col_obj_upper in columnas_actuales:
                            nombre_col_real = columnas_actuales[col_obj_upper]
                            valor_nuevo = fila[nombre_col_real]
                            if pd.notna(valor_nuevo):
                                valor_texto = str(valor_nuevo).strip()
                                if valor_texto not in ["", "-----", "nan"]:
                                    datos_maestros[rut_limpio][col_obj] = valor_nuevo
                    
                    if tiene_control_asistencia:
                        valor_firma = fila[columna_asistencia_real]
                        datos_maestros[rut_limpio][etiqueta_asistencia] = 1 if pd.notna(valor_firma) and str(valor_firma).strip() != "" else 0
                    else:
                        datos_maestros[rut_limpio][etiqueta_asistencia] = pd.NA
                        
            for rut_limpio in datos_maestros:
                if rut_limpio in ruts_ultima_nomina:
                    datos_maestros[rut_limpio]['retirado'] = 0
                else:
                    datos_maestros[rut_limpio]['retirado'] = 1
                    
            lista_final = list(datos_maestros.values())
            if not lista_final:
                st.error("No se extrajeron datos válidos. Revisa la estructura de los archivos.")
            else:
                matriz_consolidada = pd.DataFrame(lista_final)
                orden_columnas = columnas_objetivo + ['retirado'] + etiquetas_actividades
                columnas_disponibles = [col for col in orden_columnas if col in matriz_consolidada.columns]
                matriz_consolidada = matriz_consolidada[columnas_disponibles]
                
                matriz_consolidada['retirado'] = matriz_consolidada['retirado'].astype('Int64')
                for col_asist in etiquetas_actividades:
                    if col_asist in matriz_consolidada.columns:
                        matriz_consolidada[col_asist] = matriz_consolidada[col_asist].astype('Int64')
                        
                matriz_consolidada = matriz_consolidada.sort_values(by=['Primer Apellido', 'Nombres'], na_position='last')
                
                for columna in matriz_consolidada.columns:
                    if columna not in etiquetas_actividades and columna != 'retirado':
                        matriz_consolidada[columna] = matriz_consolidada[columna].apply(
                            lambda valor: valor.strftime('%d/%m/%Y') if isinstance(valor, pd.Timestamp) else (str(valor) if pd.notna(valor) and str(valor).strip() != "nan" else "")
                        )
                
                buffer_memoria = io.BytesIO()
                with pd.ExcelWriter(buffer_memoria, engine='openpyxl') as escritor:
                    matriz_consolidada.to_excel(escritor, index=False, sheet_name='Nómina Consolidada')
                    hoja = escritor.sheets['Nómina Consolidada']
                    
                    relleno_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    fuente_cabecera = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                    alineacion_cabecera = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    hoja.row_dimensions[1].height = 28
                    for celda in hoja[1]:
                        celda.fill = relleno_azul
                        celda.font = fuente_cabecera
                        celda.alignment = alineacion_cabecera
                    
                    for columna in hoja.columns:
                        largo_maximo = 0
                        for celda in columna:
                            val = str(celda.value or '')
                            if len(val) > largo_maximo:
                                largo_maximo = len(val)
                        letra_col = get_column_letter(columna[0].column)
                        hoja.column_dimensions[letra_col].width = max(largo_maximo + 3, 12)
                
                st.success("Datos unificados correctamente.")
                st.write("Vista previa de la matriz:")
                st.dataframe(matriz_consolidada)
                
                st.download_button(
                    label="Descargar archivo final PIS (.xlsx)",
                    data=buffer_memoria.getvalue(),
                    file_name="nomina_final_PIS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

with pestana_etapas:
    # --- MÓDULO 2: VALIDADOR DE VIGENCIA POR ETAPAS ---
    st.title("Validador de Vigencia por Etapas")
    st.write("Sube el archivo maestro (REVISIÓN FAMILIAS) para detectar las etapas, y luego cruza los datos con las nóminas actualizadas para determinar quién se mantiene vigente y quién fue retirado.")
    
    archivo_maestro = st.file_uploader("1. Sube el Archivo Maestro (Excel con múltiples hojas)", type=['xlsx'], key="up_maestro")
    
    if archivo_maestro:
        xls_maestro = pd.ExcelFile(archivo_maestro)
        hojas_detectadas = xls_maestro.sheet_names
        
        st.markdown("### 2. Mapeo de Nóminas Actualizadas")
        st.write("El sistema detectó las siguientes etapas en tu archivo maestro. Sube la nómina correspondiente a cada una:")
        
        archivos_actualizacion = {}
        for hoja in hojas_detectadas:
            archivos_actualizacion[hoja] = st.file_uploader(f"Nómina actualizada para: **{hoja}**", type=['xlsx'], key=f"up_{hoja}")
            
        st.divider()
        
        if st.button("Procesar Cruce de Vigencia", type="primary"):
            if not any(archivos_actualizacion.values()):
                st.warning("Debes asignar al menos un archivo actualizado para realizar la validación.")
            else:
                resultados_procesados = {}
                resumen_metricas = []
                
                # Lista blanca estricta de columnas para el archivo de salida
                columnas_maestro_permitidas = ["N°", "Rut", "Dígito", "Nombres", "Apellido Paterno", "Apellido Materno", "Retirado"]
                
                def construir_identificador_robusto(df_objetivo):
                    rut_base = df_objetivo['Rut'].astype(str).str.split('.').str[0].str.replace(r'\D', '', regex=True)
                    digito = df_objetivo['Dígito'].astype(str).str.upper().str.replace(r'[^0-9K]', '', regex=True)
                    return rut_base + digito
                
                for hoja in hojas_detectadas:
                    df_maestro = pd.read_excel(archivo_maestro, sheet_name=hoja, header=1)
                    df_maestro.columns = df_maestro.columns.str.strip()
                    
                    if 'Rut' not in df_maestro.columns or 'Dígito' not in df_maestro.columns:
                        st.error(f"Fallo estructural: La hoja '{hoja}' no contiene las columnas separadas 'Rut' y 'Dígito'. Se omitirá de la matriz final.")
                        continue
                        
                    df_maestro['llave_rut'] = construir_identificador_robusto(df_maestro)
                    
                    archivo_update = archivos_actualizacion[hoja]
                    if archivo_update is not None:
                        df_actualizado = pd.read_excel(archivo_update, header=0)
                        df_actualizado.columns = df_actualizado.columns.str.strip()
                        
                        df_actualizado['llave_rut'] = construir_identificador_robusto(df_actualizado)
                        ruts_vigentes = set(df_actualizado['llave_rut'].dropna())
                        
                        df_maestro['Retirado'] = (~df_maestro['llave_rut'].isin(ruts_vigentes)).astype(int)
                        
                        total_familias = len(df_maestro)
                        cantidad_vigentes = (df_maestro['Retirado'] == 0).sum()
                        cantidad_retirados = (df_maestro['Retirado'] == 1).sum()
                        
                        resumen_metricas.append({
                            "Etapa / Hoja": hoja,
                            "Total Maestro": total_familias,
                            "Vigentes": cantidad_vigentes,
                            "Retirados": cantidad_retirados
                        })
                    else:
                        df_maestro['Retirado'] = pd.NA
                        resumen_metricas.append({
                            "Etapa / Hoja": hoja,
                            "Total Maestro": len(df_maestro),
                            "Vigentes": "No evaluado",
                            "Retirados": "No evaluado"
                        })
                    
                    # Filtro restrictivo de limpieza: intersectar las columnas existentes con la lista blanca
                    columnas_finales = [col for col in columnas_maestro_permitidas if col in df_maestro.columns]
                    # Asegurar la retención de la columna generada lógicamente
                    if 'Retirado' not in columnas_finales and 'Retirado' in df_maestro.columns:
                        columnas_finales.append('Retirado')
                        
                    # Sobrescribir el dataframe eliminando el resto (direcciones, teléfonos, etc.)
                    df_maestro = df_maestro[columnas_finales]
                    
                    resultados_procesados[hoja] = df_maestro
                
                if resumen_metricas:
                    st.success("Cruce de datos completado.")
                    st.markdown("### Resumen Operativo")
                    st.dataframe(pd.DataFrame(resumen_metricas), hide_index=True)
                
                buffer_memoria = io.BytesIO()
                with pd.ExcelWriter(buffer_memoria, engine='openpyxl') as escritor:
                    for nombre_hoja, df_resultado in resultados_procesados.items():
                        nombre_seguro = nombre_hoja[:31]
                        df_resultado.to_excel(escritor, index=False, sheet_name=nombre_seguro)
                        hoja_excel = escritor.sheets[nombre_seguro]
                        
                        relleno_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                        fuente_cabecera = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                        alineacion_cabecera = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        hoja_excel.row_dimensions[1].height = 28
                        for celda in hoja_excel[1]:
                            celda.fill = relleno_azul
                            celda.font = fuente_cabecera
                            celda.alignment = alineacion_cabecera
                        
                        for columna in hoja_excel.columns:
                            largo_maximo = 0
                            for celda in columna:
                                val = str(celda.value or '')
                                if len(val) > largo_maximo:
                                    largo_maximo = len(val)
                            letra_col = get_column_letter(columna[0].column)
                            hoja_excel.column_dimensions[letra_col].width = min(max(largo_maximo + 3, 12), 45)
                
                st.download_button(
                    label="Descargar Maestro Limpio y Actualizado (.xlsx)",
                    data=buffer_memoria.getvalue(),
                    file_name="maestro_familias_actualizado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )